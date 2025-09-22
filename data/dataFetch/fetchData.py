from drive import get_data
import openpyxl
import io
import os
import itertools
import pandas as pd

file = get_data()
directory = None
make_version = input("Make new version? (y/n): ")
if make_version.lower() == "y":
    version = input("Version name (e.g. V1): ")
    directory = f"../{version}/"
# Abrir el archivo Excel
wb = openpyxl.load_workbook(filename=openpyxl.reader.excel.BytesIO(file), data_only=True)

# Seleccionar la hoja de "Estudiantes", "Seguimiento" y "Anteproyectos"
ws_tracking = wb["Seguimiento"]
ws_projects = wb["Anteproyectos"]

# Eliminar las columnas 3 y 4 (A, C y D) de la hoja "Seguimiento"
ws_tracking.delete_cols(4)
ws_tracking.delete_cols(3)

# Cambiar la columna 1 de la hoja "Seguimiento" por un identificador único
id_generator = itertools.count()
for row in ws_tracking.iter_rows(min_row=2, max_col=1):
    for cell in row:
        cell.value = next(id_generator)
# Cambiar el nombre de la columna A por "ID"
ws_tracking.cell(row=1, column=1, value="ID")

# Eliminar las columnas C a S de la hoja "Anteproyectos"
ws_projects.delete_cols(3, 17)

# Convertir las hojas a JSON distintos

df_tracking = pd.DataFrame(ws_tracking.values)
df_projects = pd.DataFrame(ws_projects.values)
# Asignar la primera fila como encabezado
df_tracking.columns = df_tracking.iloc[0]
df_tracking = df_tracking[1:]
df_projects.columns = df_projects.iloc[0]
df_projects = df_projects[1:]

# Eliminar columnas con nombre None
df_tracking = df_tracking.loc[:, df_tracking.columns.notna()]
df_projects = df_projects.loc[:, df_projects.columns.notna()]

# Guardar los DataFrames como archivos JSON. Si directory es None, solo en ../data/Latest. Sino, también en directory
if directory:
    # Crear el directorio si no existe
    if not os.path.exists(directory):
        os.makedirs(directory)
    df_tracking.to_json(f"{directory}/tracking.json", orient="records", force_ascii=False, indent=4)
    df_projects.to_json(f"{directory}/projects.json", orient="records", force_ascii=False, indent=4)
df_tracking.to_json("../Latest/tracking.json", orient="records", force_ascii=False, indent=4)
df_projects.to_json("../Latest/projects.json", orient="records", force_ascii=False, indent=4)

